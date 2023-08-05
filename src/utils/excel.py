from openpyxl import Workbook, load_workbook
import pandas as pd

class ExcelWork:
    def __init__(self, path, today) -> None:
        self.path = path
        self.today = today
        try:
            self.wb = load_workbook(path)
        except FileNotFoundError:
            self.wb = Workbook()


    def get_last_row(self, sheetname):
        ws = self.wb[sheetname]
        # print(len(list(ws.rows)))
        # rows = 1
        # for ro in ws:
        #     if any(col.value for col in ro):
        #         rows += 1
        # return rows
        return len(ws["B"])

    def reopen(self):
        self.wb = load_workbook(self.path)

    def add_df( self, sheetname, df, startrow, startcol, with_header=True, cols=None, extrainfo={}):
        cols = cols or df.columns
        if startrow == "end":
            startrow = self.get_last_row(sheetname) + 1
        with pd.ExcelWriter( self.path, engine="openpyxl", mode="a", if_sheet_exists="overlay", engine_kwargs={},) as writer:
            df.to_excel( excel_writer=writer, sheet_name=sheetname, float_format="%.2f", columns=cols,
                startrow=startrow, startcol=startcol, header=with_header, index=False,)
        
        self.reopen()
        ws = self.wb[sheetname]
        r, c = self.get_last_row(sheetname) + 1, startcol + len(df.columns)
        for key, info in extrainfo.items():
            c1 = ws.cell(r, c)
            c1.value = key
            c2 = ws.cell(r, c + 1)
            c2.value = info
            r += 1
        self.save()

    def save(self):
        self.wb.save(self.path)

    def delete_sheet(self, sheet_name):
        del self.wb[sheet_name]
        self.save()
        self.reopen()

    def clean(self, sheet_name):
        self.wb[sheet_name].delete_cols(1, 50)
        self.wb[sheet_name].delete_rows(1, 1000)