import os
import datetime
import pandas as pd
import dotenv
from src.connectFlattrade import initialize
from src.logger import get_logger
from openpyxl import Workbook, load_workbook

from src.utils.excel import ExcelWork

dotenv.load_dotenv()
logger = get_logger(filename="./log")

def gen_xlsx(today):
    today_str = today.strftime("%d-%m-%Y")
    strat_momentum_path = os.path.join("apidata", "momentum", today_str)
    os.makedirs(strat_momentum_path, exist_ok=True)

    # ft_data = FTDataService(logger=logger, api=api, path="./apidata")

    report_path = os.path.join(strat_momentum_path, f"report_{today_str}.csv")
    report_path_prettify = os.path.join(
        strat_momentum_path, f"report_{today_str}_cleaned.csv"
    )
    df = pd.read_csv(
        report_path,
        usecols=[
            "ordtime",
            "sym",
            "tsym",
            "token",
            "instrument",
            "lotsize",
            "avgprice",
            "qty",
            "trantype",
            "totalprice",
            "netprice",
        ],
    )
    netprofit_eq = df[df['instrument'] == 'Eq']['netprice'].sum()
    netprofit_opt = df[df['instrument'] == 'OPTSTK']['netprice'].sum()
    df.to_csv(report_path_prettify, index=None)
    df.set_index("ordtime", drop=True)

    excel_path = os.path.join(strat_momentum_path, f"report_{today_str}.xlsx")
    sheet_name = today_str

    excel = ExcelWork(excel_path, today)
    if sheet_name in excel.wb.sheetnames:
        ws = excel.wb[sheet_name]
    else:
        ws = excel.wb.create_sheet(sheet_name)
    excel.clean(sheet_name)
    def prepare_sheet(ws, today, profits):
        for col in list("ABCDEFGHIJKLMNOPQRTSUVWXYZ"):
            ws.column_dimensions[col].bestFit = True
        ws["B1"] = "After trade report"
        ws["B3"] = "Date"
        ws["C3"] = today.strftime("%a, %d %b %y")
        ws["B4"] = "Net PnL"
        ws["B5"] = "Equity PnL"
        ws["B6"] = "FnO PnL"
        ws['C4'] = profits['eq'] + profits['opt']
        ws['C5'] = profits['eq']
        ws['C6'] = profits['opt']
    prepare_sheet(ws, today, profits={'eq':netprofit_eq, 'opt':netprofit_opt})
    excel.wb.save(excel_path)
    DATA_START_COL = 1
    excel.add_df(sheet_name, df, 8, startcol=DATA_START_COL)

    gp = df.groupby(["sym", "tsym"])
    df_list = []
    for s in df.sym.unique():
        for ts in df[df["sym"] == s].tsym.unique():
            df_list.append(gp.get_group((s, ts)))

    for df in df_list:
        profit = 0
        for _, row in df.iterrows():
            profit = df['netprice'].sum()
        excel.add_df(
            sheetname=sheet_name,
            df=df,
            startrow="end",
            startcol=DATA_START_COL,
            with_header=False,
            extrainfo={"Profit :": profit},
        )
    # import sys
    # sys.exit(0)
st = datetime.datetime(2023, 7, 28)
et = datetime.datetime(2023, 8, 4)

# et = datetime.datetime.now().replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
days = pd.date_range(start=st, end=et, freq='B')
excel_path = "./apidata/momentum/reports/onlyFnO.xlsx"
excel = ExcelWork(excel_path, today)
for today in days:
    today = today.to_pydatetime()
    gen_xlsx(today)